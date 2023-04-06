import openpyxl

def write_to_excel(filename, sheetname, data):
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
    except FileNotFoundError:
        # Handle file not found error, e.g. create a new sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet.title = "Student Info"
        sheet['A1'] = "Student ID"
        sheet['B1'] = "Student Name"
        sheet['C1'] = "Site"
        sheet['D1'] = "Teacher"
        sheet['E1'] = "Grade"
        sheet['F1'] = "Asset ID"

    # Check if the value at index 0 already exists in the sheet
    current_row = sheet.max_row + 1
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True)):
        if row[0] == data[0]:
            current_row = i+2
            break

    # If the value exists, update the row    
    for i, value in enumerate(data):
        sheet.cell(row=current_row, column=i+1, value=value)
    
    workbook.save(filename)
